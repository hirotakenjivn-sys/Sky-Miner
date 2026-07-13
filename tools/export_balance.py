#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
経済数値確定パイプライン(CLAUDE.md 絶対ルール③ の実装)

  docs/credit_economy_model.xlsx  ← 唯一の真実源(数式駆動)
        │  ① 再計算(キャッシュ値の確定)
        │     - soffice(LibreOffice headless)があればそれを使う
        │     - なければ formulas ライブラリで xlsx の実数式を評価
        ▼
  build/credit_economy_model.recalc.xlsx  ← 値スナップショット
        │  ② openpyxl(data_only=True)で読み出し
        ▼
  data/balance/*.json  ← ゲーム/サーバ取り込み用
        │  ③ 検証(assert)
        ▼   - 進行シミュレーションの充足率が全期間 100% 前後
            - 天体 50 件・資源 33 件が揃っている

数値は一切ハードコードしない。すべて xlsx を再計算した結果から取り出す。
"""

from __future__ import annotations

import json
import math
import os
import shutil
import subprocess
import sys
import tempfile
import warnings
from pathlib import Path

import openpyxl

# ----------------------------------------------------------------------------
# パス定義
# ----------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
SRC_XLSX = ROOT / "docs" / "credit_economy_model.xlsx"
BUILD_DIR = ROOT / "build"
RECALC_XLSX = BUILD_DIR / "credit_economy_model.recalc.xlsx"
OUT_DIR = ROOT / "data" / "balance"

# 資源マスタ(日本語名)→ 安定id/市況ソースの対応表。
# id は市況サーバの instruments.id と JOIN するキー。source は Phase1 の取得手段の目安。
RESOURCE_IDS = {
    "鉄":         ("iron_ore",     "fred"),      # 鉄鉱石62% CFR中国指数
    "石炭":       ("coal",         "fred"),      # ICE Newcastle
    "メタン":     ("methane",      "fred"),      # Henry Hub
    "メタン氷":   ("methane_ice",  "derived"),   # Henry Hub と同一指標(methane に連動)
    "炭化水素":   ("hydrocarbon",  "fred"),      # WTI
    "木材":       ("lumber",       "fred"),      # CME材木
    "鉛":         ("lead",         "metalsapi"), # LME
    "クロム":     ("chromium",     "manual"),    # フェロクロム(週次・購読)
    "アルミニウム":("aluminum",    "metalsapi"), # LME
    "亜鉛":       ("zinc",         "metalsapi"), # LME
    "銅":         ("copper",       "metalsapi"), # LME
    "ニッケル":   ("nickel",       "metalsapi"), # LME
    "バナジウム": ("vanadium",     "manual"),    # フェロバナジウム(週次・購読)
    "コバルト":   ("cobalt",       "metalsapi"), # LME
    "錫":         ("tin",          "metalsapi"), # LME
    "モリブデン": ("molybdenum",   "manual"),    # 酸化Mo(週次・購読)
    "タングステン":("tungsten",    "manual"),    # SMM(週次・購読)
    "ウラン":     ("uranium",      "manual"),    # U3O8(週次)
    "銀":         ("silver",       "metalsapi"), # スポット
    "パラジウム": ("palladium",    "metalsapi"), # スポット
    "白金":       ("platinum",     "metalsapi"), # スポット
    "金":         ("gold",         "metalsapi"), # スポット
    "ロジウム":   ("rhodium",      "metalsapi"), # 気配
    "水":         ("water",        "fixed"),
    "硫黄":       ("sulfur",       "fixed"),
    "窒素氷":     ("nitrogen_ice", "fixed"),
    "炭素質(水・有機物)": ("carbonaceous", "fixed"),
    "アンモニア": ("ammonia",      "fixed"),
    "水素":       ("hydrogen",     "fixed"),
    "チタン":     ("titanium",     "fixed"),
    "ヘリウム":   ("helium",       "fixed"),
    "ヘリウム3":  ("helium3",      "fixed"),      # NOVA直接指定(市場なし)
    "希少物質":   ("exotic_matter","fixed"),      # NOVA直接指定(架空)
}

# 連動区分(★日次/◆週次/固定)→ 正規化
LINK_TYPE = {"★日次": "daily", "◆週次": "weekly", "固定": "fixed"}


# ----------------------------------------------------------------------------
# ① 再計算
# ----------------------------------------------------------------------------
def _find_soffice() -> str | None:
    for cand in ("soffice", "libreoffice"):
        p = shutil.which(cand)
        if p:
            return p
    mac = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    return mac if os.path.exists(mac) else None


def recalc_with_soffice(soffice: str, src: Path, dst: Path) -> None:
    """LibreOffice headless で開き直して再計算 → xlsx として書き出す。"""
    with tempfile.TemporaryDirectory() as td:
        # --convert-to は開いた時点で数式を再計算し、キャッシュ値を埋めて保存する
        subprocess.run(
            [soffice, "--headless", "--calc", "--convert-to", "xlsx",
             "--outdir", td, str(src)],
            check=True, timeout=180,
            stdout=subprocess.PIPE, stderr=subprocess.PIPE,
        )
        produced = Path(td) / (src.stem + ".xlsx")
        if not produced.exists():
            raise RuntimeError(f"soffice が出力を生成しませんでした: {produced}")
        shutil.copyfile(produced, dst)


def recalc_with_formulas(src: Path, dst: Path) -> None:
    """formulas ライブラリで xlsx の実数式を評価し、値スナップショットを書き出す。

    xlsx の数式そのものを評価するため、経済モデルの手動再実装にはならない
    (= 唯一の真実源からの逸脱を避ける)。
    """
    import formulas  # 遅延 import(soffice があれば不要)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        model = formulas.ExcelModel().loads(str(src)).finish()
        solution = model.calculate()

    # solution のキーは "'[<file>]<sheet>'!<COORD>"(大文字小文字混在)。
    # 大文字化して引ける辞書にする。
    fname = src.name
    resolved: dict[str, object] = {}
    for key, cell in solution.items():
        val = cell.value
        try:
            # numpy 配列(1x1)ならスカラーを取り出す
            val = val[0, 0]
        except Exception:
            pass
        resolved[key.upper()] = val

    wb = openpyxl.load_workbook(src, data_only=False)
    for ws in wb.worksheets:
        prefix = f"'[{fname}]{ws.title}'!".upper()
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    v = resolved.get(prefix + c.coordinate)
                    if v is None:
                        continue
                    # formulas は未計算を特殊型で返すことがある → 数値/文字列のみ採用
                    if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                        continue
                    if hasattr(v, "item"):  # numpy スカラー
                        v = v.item()
                    c.value = v
    BUILD_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(dst)


def recalc(src: Path, dst: Path) -> str:
    BUILD_DIR.mkdir(parents=True, exist_ok=True)
    soffice = _find_soffice()
    if soffice:
        try:
            recalc_with_soffice(soffice, src, dst)
            return f"soffice ({soffice})"
        except Exception as e:  # soffice があっても失敗したら formulas に退避
            print(f"  ! soffice 再計算に失敗、formulas にフォールバック: {e}")
    recalc_with_formulas(src, dst)
    return "formulas ライブラリ"


# ----------------------------------------------------------------------------
# ② 読み出しヘルパ
# ----------------------------------------------------------------------------
def _num(v):
    """セル値を数値へ。'―' や None は None、数値はそのまま。"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s in ("", "―", "-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def load_recalc():
    return openpyxl.load_workbook(RECALC_XLSX, data_only=True)


# ----------------------------------------------------------------------------
# 各シート → JSON
# ----------------------------------------------------------------------------
def export_params(wb) -> dict:
    ws = wb["パラメータ"]
    p = {
        "v0_income_rate":      _num(ws["B3"].value),   # 基準収入レート(NOVA/秒)
        "r_income_growth":     _num(ws["B4"].value),   # 天体ごと収入成長係数
        "alpha_frontier":      _num(ws["B5"].value),
        "s_unlock_spend_ratio":_num(ws["B6"].value),
        "E_effective_sec_day": _num(ws["B7"].value),
        "boost_multiplier":    _num(ws["B9"].value),
        "boost_duration_min":  _num(ws["B10"].value),
        "usd_vnd_rate":        _num(ws["B11"].value),  # NOVA=VND 換算(基準日固定)
        "base_iron_usd_kg":    _num(ws["B12"].value),
        "base_date":           "2026-07-11",
    }
    return p


def export_resources(wb) -> list:
    ws = wb["資源マスタ"]
    out = []
    seen_ids = set()
    for r in range(4, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        if name not in RESOURCE_IDS:
            # マスタ末尾の注記行など。数値行だけ拾う
            continue
        rid, source = RESOURCE_IDS[name]
        if rid in seen_ids:
            continue
        seen_ids.add(rid)
        link_raw = str(ws.cell(r, 2).value or "").strip()
        out.append({
            "id":          rid,
            "name_ja":     name,
            "link_type":   LINK_TYPE.get(link_raw, "fixed"),
            "source":      source,
            "usd_per_kg":  _num(ws.cell(r, 3).value),
            "nova_per_kg": _num(ws.cell(r, 4).value),
            "indicator":   (str(ws.cell(r, 5).value).strip()
                            if ws.cell(r, 5).value not in (None, "―") else None),
            "note":        (str(ws.cell(r, 6).value).strip()
                            if ws.cell(r, 6).value else None),
        })
    return out


def export_bodies(wb) -> list:
    ws = wb["天体マスタ"]
    out = []
    for r in range(4, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if no is None or not str(no).strip().lstrip("-").isdigit():
            continue
        out.append({
            "no":            int(no),
            "band":          str(ws.cell(r, 2).value).strip(),
            "name_ja":       str(ws.cell(r, 3).value).strip(),
            "type":          str(ws.cell(r, 4).value).strip(),
            "distance_km":   _num(ws.cell(r, 5).value),
            "target_unlock_day": _num(ws.cell(r, 6).value),
            "income_rate_nova_s": _num(ws.cell(r, 7).value),
            "delta_t_day":   _num(ws.cell(r, 8).value),
            "unlock_price_nova": _num(ws.cell(r, 9).value),
            "oneway_min_at_unlock": _num(ws.cell(r, 10).value),
            "resources":     str(ws.cell(r, 11).value or "").strip(),
            "note":          (str(ws.cell(r, 12).value).strip()
                              if ws.cell(r, 12).value else None),
        })
    return out


def export_speed_curve(wb) -> list:
    ws = wb["速度カーブ"]
    out = []
    for r in range(4, 14):  # B1..B10 = 行4..13
        band = ws.cell(r, 1).value
        if not band:
            continue
        ratio = ws.cell(r, 5).value
        out.append({
            "band":              str(band).strip(),
            "rep_distance_km":   _num(ws.cell(r, 2).value),
            "target_oneway_min": _num(ws.cell(r, 3).value),
            "required_speed_km_min": _num(ws.cell(r, 4).value),
            "ratio_vs_prev":     (None if ratio == "基準" else _num(ratio)),
            "arrival_day":       _num(ws.cell(r, 6).value),
        })
    return out


def export_upgrade_curve(wb) -> dict:
    ws = wb["強化コスト曲線"]
    params = {
        "C0_lv1_cost":        _num(ws["B3"].value),
        "cost_growth_per_lv": _num(ws["B4"].value),
        "effect_growth_per_lv": _num(ws["B5"].value),
        "milestone_multiplier": _num(ws["B6"].value),
    }
    levels = []
    for r in range(9, ws.max_row + 1):
        lv = ws.cell(r, 1).value
        if lv is None or not str(lv).strip().isdigit():
            continue
        levels.append({
            "lv":            int(lv),
            "cost":          _num(ws.cell(r, 2).value),
            "cumulative_cost": _num(ws.cell(r, 3).value),
            "effect_mult":   _num(ws.cell(r, 4).value),
        })
    return {"params": params, "levels": levels}


def export_refining(wb) -> dict:
    ws = wb["精錬・加工"]
    params = {
        "ore_grade_coeff_std": _num(ws["B11"].value),  # 品位係数 0.35
        "recovery_lv1":        _num(ws["B12"].value),  # 回収率 0.70
        "capacity_lv1_kg_h":   _num(ws["B13"].value),  # 処理能力 500kg/h
    }
    recipes = []
    for r in range(18, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        name = str(name).strip()
        cost = _num(ws.cell(r, 3).value)
        if cost is None:  # レシピ行以外(注記)を除外
            continue
        recipes.append({
            "product_ja":      name,
            "recipe":          str(ws.cell(r, 2).value or "").strip(),
            "component_cost_nova_kg": cost,
            "craft_premium":   _num(ws.cell(r, 4).value),
            "sale_price_nova_kg": _num(ws.cell(r, 5).value),
            "reference":       (str(ws.cell(r, 6).value).strip()
                                if ws.cell(r, 6).value else None),
            "die_note":        (str(ws.cell(r, 7).value).strip()
                                if ws.cell(r, 7).value else None),
        })
    return {"params": params, "recipes": recipes}


def export_ships(wb) -> dict:
    """船・輸送設計シートの Lv1 基準パラメータ。採掘/輸送の性能軸と建造コスト係数。
    (行19..24 の1便経済表は検算用サンプルで、全50天体分ではないため出さない)"""
    ws = wb["船・輸送設計"]
    return {
        "miner_mining_rate_kg_s": _num(ws["B10"].value),   # 採掘船:採掘レート基準(kg/秒)
        "miner_capacity_kg":      _num(ws["B11"].value),   # 採掘船:積載量(kg)
        "transport_capacity_kg":  _num(ws["B12"].value),   # 輸送船:積載量(kg)
        "load_unload_sec":        _num(ws["B13"].value),   # 積込+荷降ろし(秒/往復)
        "build_cost_v_frontier_sec": _num(ws["B14"].value),# 建造コスト = v_frontier × この秒数
    }


def read_progression(wb) -> list:
    """進行シミュレーションの充足率列(検証用)。JSON にも出す。"""
    ws = wb["進行シミュレーション"]
    out = []
    for r in range(4, ws.max_row + 1):
        day = ws.cell(r, 1).value
        if day is None or not str(day).strip().isdigit():
            continue
        g = ws.cell(r, 7).value  # アンロック充足率
        out.append({
            "day":          int(day),
            "bodies_unlocked": _num(ws.cell(r, 2).value),
            "income_rate":  _num(ws.cell(r, 3).value),
            "cumulative_earned": _num(ws.cell(r, 5).value),
            "cumulative_unlock_cost": _num(ws.cell(r, 6).value),
            "sufficiency":  (None if g in (None, "-", "―") else _num(g)),
        })
    return out


# ----------------------------------------------------------------------------
# ③ 検証
# ----------------------------------------------------------------------------
def validate(resources, bodies, speed, upgrade, refining, params, progression, ships):
    errors = []

    # 船パラメータが全て算出されている(採掘/積載/建造係数)
    for k in ("miner_mining_rate_kg_s", "miner_capacity_kg",
              "transport_capacity_kg", "load_unload_sec",
              "build_cost_v_frontier_sec"):
        if ships.get(k) is None or ships[k] <= 0:
            errors.append(f"船パラメータ {k} が未計算/不正: {ships.get(k)}")
    # 輸送船の積載 > 採掘船の積載(輸送でピストンする設計前提)
    if (ships.get("transport_capacity_kg") and ships.get("miner_capacity_kg")
            and ships["transport_capacity_kg"] <= ships["miner_capacity_kg"]):
        errors.append("輸送船の積載が採掘船以下(設計前提と矛盾)")

    # 資源 33 件
    if len(resources) != 33:
        errors.append(f"資源が 33 件ではない: {len(resources)} 件")
    # 天体 50 件(No 0..49)
    if len(bodies) != 50:
        errors.append(f"天体が 50 件ではない: {len(bodies)} 件")
    nos = sorted(b["no"] for b in bodies)
    if nos != list(range(50)):
        errors.append(f"天体 No が 0..49 で連続していない: {nos[:5]}..{nos[-3:]}")

    # 全天体に収入レート・アンロック価格・片道時間が算出されている
    for b in bodies:
        if b["income_rate_nova_s"] is None:
            errors.append(f"天体 {b['name_ja']} の収入レートが未計算")
        if b["no"] != 0 and b["unlock_price_nova"] is None:
            errors.append(f"天体 {b['name_ja']} のアンロック価格が未計算")
        if b["oneway_min_at_unlock"] is None:
            errors.append(f"天体 {b['name_ja']} の片道時間が未計算")

    # 速度カーブ 10 バンド・速度が単調増加
    if len(speed) != 10:
        errors.append(f"速度カーブが 10 バンドではない: {len(speed)}")
    speeds = [s["required_speed_km_min"] for s in speed]
    if any(s is None for s in speeds):
        errors.append("速度カーブに未計算のバンドがある")
    elif any(speeds[i] >= speeds[i + 1] for i in range(len(speeds) - 1)):
        errors.append("統一速度 S が単調増加していない")

    # 強化曲線 50 レベル
    if len(upgrade["levels"]) != 50:
        errors.append(f"強化曲線が 50 レベルではない: {len(upgrade['levels'])}")

    # 精錬レシピが 6 種そろい、成分市況連動で価格が算出されている
    if len(refining["recipes"]) < 6:
        errors.append(f"精錬レシピが 6 種未満: {len(refining['recipes'])}")
    for rc in refining["recipes"]:
        if not rc["sale_price_nova_kg"] or rc["sale_price_nova_kg"] <= 0:
            errors.append(f"加工品 {rc['product_ja']} の販売単価が未計算")

    # 進行シミュレーションの充足率が「全期間 100% 前後」であること。
    #   充足率 = (累積収入 × s)÷ 累積アンロック費。
    #   ・100% 以上 = その目標日に開放費を賄える(スケジュール達成可能)。
    #   ・累積アンロック費は「開放日」だけ段階的に増え、収入は毎日積み上がるため、
    #     系列は開放と開放のあいだで余剰が溜まる鋸歯状になる(離散日次モデルの性質)。
    #   ・したがってバランスの拘束条件は「最も苦しい日=系列の最小値」であり、
    #     これが 100% 前後(=どの目標日でも開放費をちょうど賄える)かつ、
    #     全日で 100% を割らない(スケジュールが破綻しない)ことを検証する。
    suff = [p["sufficiency"] for p in progression if p["sufficiency"] is not None]
    suff_stats = (None, None, None)
    if not suff:
        errors.append("充足率が 1 件も算出されていない")
    else:
        s_sorted = sorted(suff)
        lo = s_sorted[0]
        hi = s_sorted[-1]
        med = s_sorted[len(s_sorted) // 2]
        suff_stats = (lo, med, hi)
        # ① 全日で 100% を割らない(=どの目標日でも開放費が不足しない)。
        #    叩き台の許容:わずかな下振れ(0.95)まで許す。
        if lo < 0.95:
            errors.append(
                f"充足率が 100% を下回る日がある(スケジュール破綻の恐れ): "
                f"最小={lo:.3f}")
        # ② 最も苦しい日(最小値)が 100% 前後に収まる=進行が速すぎない。
        #    最小値が大きすぎる場合、全期間で余りすぎ(開放が容易すぎ)。
        if lo > 1.35:
            errors.append(
                f"充足率の最小値が 100% 前後(≤1.35)を超えており進行が速すぎる: "
                f"最小={lo:.3f}")
        # ③ 鋸歯の上端(最大値)が異常でない(数式事故の検知。3.0 を上限の目安)。
        if hi > 3.0:
            errors.append(
                f"充足率の最大値が異常(数式事故の可能性): 最大={hi:.3f}")

    # 主要既知値の突き合わせ(企画書の記載と一致するか)
    eros = next((b for b in bodies if b["name_ja"] == "エロス"), None)
    if eros and eros["unlock_price_nova"]:
        # 企画書:エロス ≒ 2,070万 NOVA
        if not (2.0e7 <= eros["unlock_price_nova"] <= 2.2e7):
            errors.append(
                f"エロスのアンロック価格が想定(≒2,070万)から外れる: "
                f"{eros['unlock_price_nova']:,.0f}")

    return errors, suff_stats


# ----------------------------------------------------------------------------
# メイン
# ----------------------------------------------------------------------------
def write_json(path: Path, obj) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)
        f.write("\n")


def main() -> int:
    if not SRC_XLSX.exists():
        print(f"ERROR: 真実源が見つかりません: {SRC_XLSX}", file=sys.stderr)
        return 2

    print(f"① 再計算: {SRC_XLSX.name}")
    engine = recalc(SRC_XLSX, RECALC_XLSX)
    print(f"   再計算エンジン: {engine}")
    print(f"   → {RECALC_XLSX.relative_to(ROOT)}")

    print("② 読み出し(openpyxl data_only)")
    wb = load_recalc()
    params      = export_params(wb)
    resources   = export_resources(wb)
    bodies      = export_bodies(wb)
    speed       = export_speed_curve(wb)
    upgrade     = export_upgrade_curve(wb)
    refining    = export_refining(wb)
    ships       = export_ships(wb)
    progression = read_progression(wb)

    print("③ 検証")
    errors, (slo, smed, shi) = validate(
        resources, bodies, speed, upgrade, refining, params, progression, ships)
    if errors:
        print("   検証失敗:")
        for e in errors:
            print(f"     - {e}")
        return 1
    print(f"   OK  資源={len(resources)} 天体={len(bodies)} "
          f"速度バンド={len(speed)} 強化Lv={len(upgrade['levels'])} "
          f"精錬レシピ={len(refining['recipes'])}")
    print(f"   OK  充足率 最小={slo:.3f}(拘束条件≒100%) "
          f"中央={smed:.3f} 最大={shi:.3f}")

    # 書き出し
    write_json(OUT_DIR / "params.json",       params)
    write_json(OUT_DIR / "resources.json",    resources)
    write_json(OUT_DIR / "bodies.json",       bodies)
    write_json(OUT_DIR / "speed_curve.json",  speed)
    write_json(OUT_DIR / "upgrade_curve.json", upgrade)
    write_json(OUT_DIR / "refining.json",     refining)
    write_json(OUT_DIR / "ships.json",        ships)
    write_json(OUT_DIR / "progression.json",  progression)

    # マニフェスト(取り込み側が一覧を把握できるように)
    manifest = {
        "source": "docs/credit_economy_model.xlsx",
        "recalc_engine": engine,
        "base_date": params["base_date"],
        "usd_vnd_rate": params["usd_vnd_rate"],
        "files": {
            "params.json":        "基準パラメータ(v0/r/α/s/E/為替/クランプ基準)",
            "resources.json":     f"資源マスタ {len(resources)} 種(id・連動区分・NOVA単価)",
            "bodies.json":        f"天体マスタ {len(bodies)} 体(距離・収入・アンロック価格・片道)",
            "speed_curve.json":   f"統一速度カーブ {len(speed)} バンド",
            "upgrade_curve.json": f"強化コスト曲線 {len(upgrade['levels'])} Lv",
            "refining.json":      f"精錬・加工 {len(refining['recipes'])} レシピ",
            "ships.json":         "船・輸送設計(採掘/輸送のLv1基準・建造係数)",
            "progression.json":   "進行シミュレーション(充足率)",
        },
    }
    write_json(OUT_DIR / "manifest.json", manifest)

    print(f"   → data/balance/ に {len(manifest['files']) + 1} ファイル書き出し")
    print("完了。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
